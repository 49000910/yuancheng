import sys, os, re, time, subprocess, base64, email, email.utils, json, datetime, urllib.parse, gc
import pythoncom
import win32com.client

import openpyxl
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QLineEdit, QPushButton, QTextEdit, 
                             QSystemTrayIcon, QMenu, QAction, QStyle, QFileDialog)
from PyQt5.QtCore import QTimer, Qt, QUrl, QEvent, QThread, pyqtSignal
from PyQt5.QtWebEngineWidgets import QWebEngineView, QWebEngineProfile

class OutlookNewMailHandler:
    app_window = None

    def OnNewMailEx(self, EntryIDCollection):
        try:
            if OutlookNewMailHandler.app_window:
                OutlookNewMailHandler.app_window.on_outlook_new_mail(EntryIDCollection)
        except Exception:
            pass
class PowerShellSyncWorker(QThread):
    finished_result = pyqtSignal(bool, int, str)

    def __init__(self, ps_cmd, timeout_sec=180):
        super().__init__()
        self.ps_cmd = ps_cmd
        self.timeout_sec = timeout_sec

    def run(self):
        try:
            ps_b64 = base64.b64encode(self.ps_cmd.encode('utf-16-le')).decode('ascii')

            proc = subprocess.run(
                ["powershell", "-NoProfile", "-EncodedCommand", ps_b64],
                capture_output=True,
                text=True,
                creationflags=0x08000000,
                timeout=self.timeout_sec
            )

            out = proc.stdout or proc.stderr or ""

            if "RESULT_SUCCESS" in out:
                m = re.search(r'COUNT:(\d+)', out)
                count = int(m.group(1)) if m else 0
                self.finished_result.emit(True, count, out)
            else:
                self.finished_result.emit(False, 0, out)

        except Exception as e:
            self.finished_result.emit(False, 0, str(e))

class OutlookMHTMaster(QWidget):
    def __init__(self):
        super().__init__()
        # --- 核心配置 ---
        self.share_dir = r'\\10.1.93.32\DT_HU_RDteam_F\视频\Z\ZOUQIU\paican'
        self.target_kw = 'EDFA' 
        self.tag_regex = r'\bEP[A-Z0-9]{9}\b' 
        self.interval_min = 30     
        self.web_refresh_sec = 2000  
        self.start_hour = 12       
        self.end_hour = 13        
        self.theme_color = "#107c10" 
        self.web_title = "EDFA排产看板"
        self.web_sub_title = "自动同步邮箱 zouqiu@huawei.com"
        self.copyright_text = "© 月入2800每天笑哈哈"
        
        # 内存优化：限制浏览器内核缓存
        QWebEngineProfile.defaultProfile().setHttpCacheType(QWebEngineProfile.MemoryHttpCache)
        QWebEngineProfile.defaultProfile().setHttpCacheMaximumSize(50 * 1024 * 1024) 

        self.last_sync_time = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
        self.link_file = os.path.join(self.share_dir, "links.txt")
        self.sync_busy = False
        self.sync_worker = None
        self.init_ui()
        self.init_tray()


        # Outlook 新邮件通知监听：24小时有效，不受定时扫描时段限制
        self.mail_event_busy = False
        self.ol_event_app = None
        self.ol_event_ns = None
        self.init_outlook_event_listener()

        QApplication.setQuitOnLastWindowClosed(False)

        self.sync_timer = QTimer(self)
        self.sync_timer.timeout.connect(self.run_cycle)

        # 启动后仍然按原逻辑检查一次，受时段限制
        QTimer.singleShot(1000, self.run_cycle)



    def init_ui(self):
        self.setWindowTitle("EDFA 看板后台 V68.0")
        self.resize(520, 850)
        layout = QVBoxLayout()
        def quick_edit(label, val, attr):
            l = QHBoxLayout(); lb = QLabel(label); lb.setFixedWidth(110); l.addWidget(lb)
            edit = QLineEdit(str(val)); setattr(self, attr, edit); l.addWidget(edit); layout.addLayout(l)
        quick_edit("📂 共享路径", self.share_dir, "ui_path")
        quick_edit("📧 邮件关键词", self.target_kw, "ui_kw")
        quick_edit("🔍 提取正则", self.tag_regex, "ui_regex")
        quick_edit("🚩 网页大标题", self.web_title, "ui_title")
        quick_edit("📝 网页备注", self.web_sub_title, "ui_subtitle")
        h1 = QHBoxLayout(); h1.addWidget(QLabel("⏱ 周期(分)")); self.ui_freq = QLineEdit(str(self.interval_min)); h1.addWidget(self.ui_freq)
        h1.addWidget(QLabel("🌐 刷新(秒)")); self.ui_web_freq = QLineEdit(str(self.web_refresh_sec)); h1.addWidget(self.ui_web_freq); layout.addLayout(h1)
        h2 = QHBoxLayout(); h2.addWidget(QLabel("🔢 抓取数量")); self.ui_count = QLineEdit(str(5)); h2.addWidget(self.ui_count)
        h2.addWidget(QLabel("⏰ 时间")); self.ui_start = QLineEdit(str(self.start_hour)); h2.addWidget(self.ui_start)
        h2.addWidget(QLabel("-")); self.ui_end = QLineEdit(str(self.end_hour)); h2.addWidget(self.ui_end); layout.addLayout(h2)
        quick_edit("🎨 主题颜色", self.theme_color, "ui_color"); quick_edit("🔒 版权内容", self.copyright_text, "ui_copy")
        
        h_btns = QVBoxLayout(); h_row1 = QHBoxLayout()
        self.btn_import_fav = QPushButton("📂 导入浏览器书签"); self.btn_import_fav.setFixedHeight(40); self.btn_import_fav.clicked.connect(self.import_bookmarks)
        self.btn_edit_links = QPushButton("📝 手动编辑链接"); self.btn_edit_links.setFixedHeight(40); self.btn_edit_links.clicked.connect(lambda: os.startfile(self.link_file))
        h_row1.addWidget(self.btn_import_fav); h_row1.addWidget(self.btn_edit_links)
        self.btn_apply = QPushButton("🚀 立即同步并刷新看板"); self.btn_apply.setFixedHeight(50); self.btn_apply.clicked.connect(self.apply_settings)
        h_btns.addLayout(h_row1); h_btns.addWidget(self.btn_apply); layout.addLayout(h_btns)
        
        self.log_area = QTextEdit(); self.log_area.setReadOnly(True); layout.addWidget(self.log_area)
        self.web = QWebEngineView(); self.setLayout(layout); self.restyle()

    def init_tray(self):
        # 創建托盤圖標
        self.tray = QSystemTrayIcon(self)
        self.tray.setIcon(self.style().standardIcon(QStyle.SP_ComputerIcon))
        tm = QMenu()
        tm.addAction("顯示主窗", self.showNormal)
        tm.addAction("徹底退出", QApplication.instance().quit)
        self.tray.setContextMenu(tm)
        self.tray.show()

    def closeEvent(self, event):
        if self.tray.isVisible():
            self.hide()
            event.ignore()
        else:
            event.accept()

    def changeEvent(self, event):
        if event.type() == QEvent.WindowStateChange and self.isMinimized():
            self.hide()
        super().changeEvent(event)
    def restyle(self):
        c = self.ui_color.text().strip() or "#107c10"
        self.setStyleSheet(f"QPushButton{{background:{c};color:white;font-weight:bold;border-radius:4px;}}")
        self.btn_edit_links.setStyleSheet("background:#555; color:white;")
        self.btn_import_fav.setStyleSheet("background:#444; color:white;")

    def add_log(self, txt):
        self.log_area.append(f"[{time.strftime('%H:%M:%S')}] {str(txt)}")



    def init_outlook_event_listener(self):
        """
        Outlook 新邮件事件监听。
        24小时监听，不受 8-18 定时扫描时段限制。
        """
        try:
            pythoncom.CoInitialize()

            OutlookNewMailHandler.app_window = self

            self.ol_event_app = win32com.client.DispatchWithEvents(
                "Outlook.Application",
                OutlookNewMailHandler
            )

            self.ol_event_ns = self.ol_event_app.GetNamespace("MAPI")

            # PyQt 主循环里泵 COM 消息，否则 Outlook 事件不会进来
            self.com_pump_timer = QTimer(self)
            self.com_pump_timer.timeout.connect(pythoncom.PumpWaitingMessages)
            self.com_pump_timer.start(500)

            self.add_log("📨 Outlook 新邮件通知监听已启动，24小时有效")

        except Exception as e:
            self.add_log(f"⚠️ Outlook 通知监听启动失败，继续使用定时扫描: {e}")

    def on_outlook_new_mail(self, entry_ids):
        """
        Outlook 收到新邮件时触发。
        只要主题包含关键词，就延迟 8 秒后立即同步，不受时段限制。
        """
        try:
            ids = str(entry_ids).split(",")

            kw = self.ui_kw.text().strip()
            if not kw:
                return

            hit = False

            for eid in ids:
                eid = eid.strip()

                if not eid:
                    continue

                try:
                    item = self.ol_event_ns.GetItemFromID(eid)
                    subject = getattr(item, "Subject", "") or ""

                    if kw.lower() in subject.lower():
                        hit = True
                        self.add_log(f"📬 新邮件通知命中关键词: {subject}")
                        break

                except Exception as e:
                    self.add_log(f"⚠️ 新邮件详情读取失败: {e}")

            if not hit:
                return

            if self.mail_event_busy:
                self.add_log("⏳ 新邮件同步已在等待/执行中，本次通知合并处理")
                return

            self.mail_event_busy = True

            self.add_log("⏱ 等待 8 秒，确保邮件正文和图片下载完整...")
            QTimer.singleShot(8000, self._run_shell_from_mail_event)

        except Exception as e:
            self.add_log(f"⚠️ 新邮件事件处理异常: {e}")

    def _run_shell_from_mail_event(self):
        """
        邮件通知触发同步。
        run_shell 现在只启动 QThread 后台同步，不会卡界面。
        """
        try:
            self.add_log("🚀 Outlook 新邮件通知触发同步，跳过时段限制")
            self.run_shell()
        finally:
            QTimer.singleShot(30000, lambda: setattr(self, "mail_event_busy", False))

    def decode_part_text(self, part):

        payload = part.get_payload(decode=True)

        if payload is None:
            return part.get_payload() or ""

        charset = part.get_content_charset() or "utf-8"

        for cs in [charset, "utf-8", "gb18030", "gbk", "big5"]:
            try:
                return payload.decode(cs, errors="ignore")
            except Exception:
                pass

        return payload.decode("utf-8", errors="ignore")


    def mht_to_single_html_content(self, mht_path):
        """
        最小改动版：
        读取 mht，把 cid 图片、Content-Location 图片转成 base64，塞进 HTML。
        返回：msg, html_content
        """
        with open(mht_path, "rb") as fp:
            msg = email.message_from_binary_file(fp)

        html_content = ""

        # 1. 取 HTML 正文
        for part in msg.walk():
            if part.get_content_type() == "text/html":
                html_content = self.decode_part_text(part)
                break

        replacements = {}

        # 2. 把 MHT 里的图片转成 data:image/base64
        for part in msg.walk():
            if part.get_content_maintype() != "image":
                continue

            data = part.get_payload(decode=True)
            if not data:
                continue

            mime = part.get_content_type() or "image/png"
            b64 = base64.b64encode(data).decode("ascii")
            data_uri = f"data:{mime};base64,{b64}"

            cid = (part.get("Content-ID") or "").strip().strip("<>")
            loc = (part.get("Content-Location") or "").strip()

            keys = []

            if cid:
                keys.append(f"cid:{cid}")
                keys.append(f"cid:{urllib.parse.quote(cid)}")
                keys.append(cid)

            if loc:
                loc_unquote = urllib.parse.unquote(loc)
                loc_base = os.path.basename(loc_unquote.split("?")[0].split("#")[0])

                keys.append(loc)
                keys.append(loc_unquote)

                if loc_base:
                    keys.append(loc_base)

            for k in keys:
                if k:
                    replacements[k] = data_uri

        def find_replacement(src):
            if not src:
                return None

            src_unquote = urllib.parse.unquote(src)
            src_base = os.path.basename(src_unquote.split("?")[0].split("#")[0])

            candidates = [src, src_unquote, src_base]

            if src.lower().startswith("cid:"):
                cid_val = src[4:]
                candidates.append(cid_val)
                candidates.append(urllib.parse.unquote(cid_val))

            for c in candidates:
                if c in replacements:
                    return replacements[c]

            return None

        # 3. 替换 src=""
        def replace_src(m):
            quote = m.group(1)
            src = m.group(2)

            if src.lower().startswith("data:image"):
                return m.group(0)

            new_src = find_replacement(src)

            if new_src:
                return f'src={quote}{new_src}{quote}'

            return m.group(0)

        html_content = re.sub(
            r'src\s*=\s*([\'"])(.*?)\1',
            replace_src,
            html_content,
            flags=re.IGNORECASE
        )

        # 4. 替换 background=""
        def replace_background(m):
            quote = m.group(1)
            src = m.group(2)

            if src.lower().startswith("data:image"):
                return m.group(0)

            new_src = find_replacement(src)

            if new_src:
                return f'background={quote}{new_src}{quote}'

            return m.group(0)

        html_content = re.sub(
            r'background\s*=\s*([\'"])(.*?)\1',
            replace_background,
            html_content,
            flags=re.IGNORECASE
        )

        # 5. 替换 CSS 里的 url(cid:xxx)
        def replace_css_url(m):
            src = m.group(2)

            if src.lower().startswith("data:image"):
                return m.group(0)

            new_src = find_replacement(src)

            if new_src:
                return f'url("{new_src}")'

            return m.group(0)

        pattern_url = "url" + re.escape("(") + r"\s*(['\"]?)(.*?)\1\s*" + re.escape(")")

        html_content = re.sub(
            pattern_url,
            replace_css_url,
            html_content,
            flags=re.IGNORECASE
        )


        return msg, html_content

    def apply_settings(self):
        self.restyle()
        self.add_log("⚙️ 配置已應用，正在檢查時段...")
        # 調用 run_cycle 進行時段判定，而不是直接運行抓取
        self.run_cycle()

    def import_bookmarks(self):
        path, _ = QFileDialog.getOpenFileName(self, "選擇瀏覽器導出的書籤HTML文件", "", "HTML Files (*.html)")
        if not path: return
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f: content = f.read()
            bookmarks = re.findall(r'<A HREF="(.*?)".*?>(.*?)</A>', content, re.IGNORECASE)
            if bookmarks:
                with open(self.link_file, 'w', encoding='utf-8') as f:
                    for url, name in bookmarks:
                        if url.startswith('http'): f.write(f"{name.strip()},{url.strip()}\n")
                self.add_log(f"✅ 成功導入 {len(bookmarks)} 個書籤")
                self.process_web()
            else: self.add_log("❌ 未發現有效書籤")
        except Exception as e: self.add_log(f"ERR: {e}")

    def run_cycle(self):
        now_h = int(time.strftime("%H"))
        try:
            s, e = int(self.ui_start.text()), int(self.ui_end.text())
            # 【時段攔截】：非活躍時段徹底不進入 run_shell，不執行任何文件操作
            if not (s <= now_h < e):
                self.add_log(f"💤 非活躍時段 ({now_h}點)，跳過同步任務")
                self.sync_timer.start(30 * 60000) 
                return
            
            # 活躍時段內執行抓取
            self.run_shell()
            self.sync_timer.start(int(self.ui_freq.text()) * 60000)
        except Exception as err:
            self.add_log(f"⏰ 調度異常: {err}")
            self.sync_timer.start(600000)

    def run_shell(self):
        d = self.ui_path.text().replace('"', '').strip()
        k = self.ui_kw.text().strip()
        try: c_num = int(self.ui_count.text())
        except: c_num = 3
        
        save_limit = 25    # 列表保存總量上限
        days_limit = 25    # 過期期限

        ps_cmd = f"""
        $ErrorActionPreference = 'Stop'
        $saveDir = "{d}"
        try {{
            $ol = New-Object -ComObject Outlook.Application
            $ns = $ol.GetNamespace("MAPI")
            
            # 1. 檢索：按設定數量抓取最新郵件
            $items = $ns.GetDefaultFolder(6).Items | Where-Object {{ 
                $_.ReceivedTime -gt (Get-Date).AddDays(-30) -and ($_.Subject -like "*{k}*") 
            }} | Sort-Object ReceivedTime -Descending | Select-Object -First {c_num}
            
                       # 先计算哪些邮件需要保存，不立刻 SaveAs，避免先占磁盘
            $pending = @()

            foreach($m in $items) {{
                $safeName = ($m.Subject -replace '[\\x00-\\x1f\\\\/:*?"<>|]', '_').Trim()

                if ([string]::IsNullOrWhiteSpace($safeName)) {{
                    $safeName = "NoSubject"
                }}

                if ($safeName.Length -gt 160) {{
                    $safeName = $safeName.Substring(0, 160)
                }}

                $target = Join-Path $saveDir "$safeName.mht"

                # 只有新邮件才加入待保存列表，但此处不保存
                if (-not (Test-Path $target) -or ($m.ReceivedTime -gt (Get-Item $target).LastWriteTime)) {{
                    $pending += [PSCustomObject]@{{
                        Mail = $m
                        Target = $target
                    }}
                }}
            }}

            $newCount = $pending.Count

            # 只有确定有新邮件要保存时，才先清理腾空间
            if ($newCount -gt 0) {{

                # A. 先删除 25 天前旧文件，index.html 不删
                $limitDate = (Get-Date).AddDays(-{days_limit})

                Get-ChildItem -Path $saveDir -File | Where-Object {{
                    ($_.Extension -in ".mht", ".html") -and
                    ($_.Name -ne "index.html") -and
                    ($_.LastWriteTime -lt $limitDate)
                }} | Remove-Item -Force -ErrorAction SilentlyContinue

                # B. 提前腾出新邮件数量的位置
                # 例如上限25，本次要保存5封，就先只保留最新20封旧文件
                $preKeep = [Math]::Max(0, {save_limit} - $newCount)

                $allBefore = @(Get-ChildItem -Path $saveDir -File | Where-Object {{
                    ($_.Extension -in ".mht", ".html") -and
                    ($_.Name -ne "index.html")
                }} | Sort-Object LastWriteTime -Descending)

                if ($allBefore.Count -gt $preKeep) {{
                    $allBefore | Select-Object -Skip $preKeep | Remove-Item -Force -ErrorAction SilentlyContinue
                }}

                # C. 清理完成后，再保存新邮件
                foreach($p in $pending) {{

                    # 如果同名旧文件还存在，先删掉，避免 SaveAs 覆盖失败，也能先释放空间
                    if (Test-Path $p.Target) {{
                        Remove-Item $p.Target -Force -ErrorAction SilentlyContinue
                    }}

                    $p.Mail.SaveAs($p.Target, 10)
                }}

                # D. 保存完成后，再校正一次，总量最多 25，index.html 不删
                $allAfter = @(Get-ChildItem -Path $saveDir -File | Where-Object {{
                    ($_.Extension -in ".mht", ".html") -and
                    ($_.Name -ne "index.html")
                }} | Sort-Object LastWriteTime -Descending)

                if ($allAfter.Count -gt {save_limit}) {{
                    $allAfter | Select-Object -Skip {save_limit} | Remove-Item -Force -ErrorAction SilentlyContinue
                }}
            }}

            Write-Host "RESULT_SUCCESS|COUNT:$newCount"

        }} catch {{ 
            Write-Host "RESULT_ERROR|$($_.Exception.Message)" 
        }} finally {{ 
            if ($ol) {{ [System.Runtime.Interopservices.Marshal]::ReleaseComObject($ol) | Out-Null }}
        }}
        """
        try:
            if self.sync_busy:
                self.add_log("⏳ 同步任务正在执行中，本次请求已跳过")
                return

            self.sync_busy = True

            self.add_log(f"🔍 后台检查中 (抓取:{c_num} / 上限:{save_limit})...")

            self.sync_worker = PowerShellSyncWorker(ps_cmd, timeout_sec=180)
            self.sync_worker.finished_result.connect(self.on_sync_finished)
            self.sync_worker.start()

        except Exception as e:
            self.sync_busy = False
            self.add_log(f"💥 启动同步任务失败: {str(e)}")

    def on_sync_finished(self, ok, count, msg):
        self.sync_busy = False

        try:
            if ok:
                if count > 0:
                    self.add_log(f"📩 发现 {count} 封新邮件，已执行保存与清理")
                else:
                    self.add_log("✅ 暂无新邮件，仅刷新看板")

                self.process_web(is_active=True)

            else:
                self.add_log(f"⚠️ 异常: {msg}")

        except Exception as e:
            self.add_log(f"💥 同步完成处理异常: {e}")

        finally:
            self.sync_worker = None

    def process_web(self, is_active=True):
        d = self.ui_path.text().strip()
        if not os.path.exists(d): return
        files = [x for x in os.listdir(d) if x.endswith('.mht')]
        files.sort(key=lambda x: os.path.getmtime(os.path.join(d, x)), reverse=True)
        mail_data = []
        # --- 内存保护：限制网页显示的邮件数量为 15 封 ---
        for f in files[:25]: 
            try:
                p_m = os.path.join(d, f)
                msg, content = self.mht_to_single_html_content(p_m)

                raw_date = msg.get('Date')
                dt_str = datetime.datetime.fromtimestamp(os.path.getmtime(p_m)).strftime('%m-%d %H:%M')

                if raw_date:
                    t_ptr = email.utils.parsedate_tz(raw_date)
                    if t_ptr:
                        dt_str = datetime.datetime.fromtimestamp(email.utils.mktime_tz(t_ptr)).strftime('%m-%d %H:%M')

                # 搜索标签时去掉 base64，避免正则扫大图导致变慢
                content_for_search = re.sub(
                    r'data:image/[^"\']+',
                    '',
                    content,
                    flags=re.IGNORECASE
                )

                tags = " ".join(list(set(re.findall(self.ui_regex.text(), content_for_search))))

                mail_data.append({
                    "title": f.replace('.mht',''),
                    "time": dt_str,
                    "body": content,
                    "raw": urllib.parse.quote(content),
                    "tags": tags
                })

            except: pass
        
        link_html = ""
        if os.path.exists(self.link_file):
            try:
                with open(self.link_file, 'r', encoding='utf-8') as f:
                    for line in f:
                        if ',' in line: n, u = line.strip().split(',', 1); link_html += f'<a href="{u.strip()}" target="_blank">🔗 {n.strip()}</a>'
            except: pass

        cal_html = ""
        for fn in os.listdir(d):
            if "2026日历" in fn and fn.lower().endswith('.xlsx'):
                try:
                    wb = openpyxl.load_workbook(os.path.join(d, fn), data_only=True); ws = wb.active; m_ranges = ws.merged_cells.ranges
                    rows_h = ""
                    for row in ws.iter_rows():
                        row_c = ""
                        for cell in row:
                            if any(cell.coordinate in m and cell.coordinate != m.start_cell.coordinate for m in m_ranges): continue
                            v, bg = "" if cell.value is None else str(cell.value), "white"
                            if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color.index != "00000000":
                                try: bg = f"#{cell.fill.start_color.rgb[2:]}"
                                except: pass
                            cs, rs = 1, 1
                            for m in m_ranges:
                                if cell.coordinate == m.start_cell.coordinate: cs = m.max_col - m.min_col + 1; rs = m.max_row - m.min_row + 1; break
                            row_c += f'<td colspan="{cs}" rowspan="{rs}" style="background:{bg}; border:1px solid #eee; padding:6px; font-size:12px; text-align:center;">{v}</td>'
                        rows_h += f"<tr>{row_c}</tr>"
                    cal_html = f'<table style="border-collapse:collapse; width:100%;">{rows_h}</table>'
                except: pass

        list_items = ""; body_sections = ""
        for i, item in enumerate(mail_data):
            list_items += f'<div class="mail-item" id="mi-{i}" onclick="showMail(\'{i}\', this)" data-tags="{item["tags"]}"><div style="display:flex; flex-direction:column; gap:4px;"><span class="mail-title">{item["title"]}</span><span class="time-badge">收件时间：{item["time"]}</span></div></div>'
            body_sections += f'<div id="mail-{i}" class="mail-body" style="display:none" data-raw="{item["raw"]}"><div class="mail-inner-zoom">{item["body"]}</div></div>'

        main_c = self.ui_color.text() or "#107c10"
        full_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
    body {{ font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif; margin: 0; display: flex; height: 100vh; background:#f3f2f1; overflow: hidden; }}
    .sidebar {{ width: 340px; background: white; border-right: 1px solid #edebe9; display: flex; flex-direction: column; position: relative; }}
    .header {{ padding: 20px 16px; background: {main_c}; color: white; }}
    .search-box {{ padding: 12px 16px; border-bottom: 1px solid #f3f2f1; position: relative; display: flex; align-items: center; }}
    .search-box input {{ width: 100%; padding: 8px 30px 8px 10px; border: 1px solid #ddd; border-radius: 6px; outline:none; }}
    .clear-btn {{ position: absolute; right: 25px; color: #ccc; cursor: pointer; font-size: 18px; display: none; }}
    .search-box input:not(:placeholder-shown) + .clear-btn {{ display: block; }}
    .mail-list {{ flex: 1; overflow-y: auto; }}
    .mail-item {{ padding: 14px 16px; border-bottom: 1px solid #f3f2f1; cursor: pointer; transition: background 0.2s; }}
    .mail-title {{ font-size: 13px; color: #333; word-break: break-all; }}
    .time-badge {{ font-size: 10px; color: white !important; background: #666; padding: 2px 10px; border-radius: 12px; white-space: nowrap; width: fit-content; align-self: flex-start; }}
    .mail-item.active {{ border-left: 5px solid {main_c}; background: #eff6ef; font-weight: bold; }}
    .mail-item.active .time-badge {{ background: {main_c}; }}
    .content {{ flex: 1; background: #f8f9fa; overflow: auto; display: flex; flex-direction: column; }}
    .mail-inner-zoom {{ padding: 25px; zoom: 0.9; background: white; margin: 20px auto; width: 92%; box-shadow: 0 4px 15px rgba(0,0,0,0.08); border-radius: 12px; }}
    .focus-row td {{ background-color: #ff3333 !important; color: white !important; font-weight: bold !important; }}
    mark {{ background: yellow; color: black; }}
    #bubble {{ position: absolute; left: 15px; bottom: 125px; right: 15px; background: rgba(33, 33, 33, 0.9); padding: 10px; border-radius: 8px; color: #00ff00; font-size:12px; text-align: center; z-index: 100; transition: opacity 0.8s; }}
    .footer {{ padding: 15px; border-top: 1px solid #f3f2f1; background: white; position: relative; }}
    .btn-group {{ display: flex; gap: 8px; margin-bottom: 10px; }}
    .btn-action {{ flex: 1; padding: 10px 5px; border: none; border-radius: 8px; background: #f3f2f1; color: #333; font-size: 12px; font-weight: 600; cursor: pointer; transition: all 0.2s; display: flex; align-items: center; justify-content: center; gap: 5px; }}
    .btn-action:hover {{ background: {main_c}; color: white; }}
    .status-bar {{ display: flex; align-items: center; justify-content: center; gap: 8px; font-size: 11px; color: #888; margin-bottom: 4px; }}
    .dot {{ width: 8px; height: 8px; background-color: {main_c}; border-radius: 50%; animation: breathe 2s infinite ease-in-out; }}
    @keyframes breathe {{ 0% {{ opacity: 0.3; transform: scale(0.9); }} 50% {{ opacity: 1; transform: scale(1.1); }} 100% {{ opacity: 0.3; transform: scale(0.9); }} }}
    .copy-text {{ font-size: 9px; color: #bbb; text-align: center; }}
    .link-flyout {{ position: absolute; bottom: 130px; left: 15px; width: 200px; background: white; box-shadow: 0 5px 20px rgba(0,0,0,0.1); display: none; border-radius: 8px; overflow-y: auto; max-height: 250px; z-index: 1001; }}
    .link-flyout a {{ display: block; padding: 10px 15px; color: #444; text-decoration: none; font-size: 12px; border-bottom: 1px solid #f9f9f9; }}
    .link-flyout a:hover {{ background: #f3f2f1; color: {main_c}; }}
    .modal {{ display: none; position: fixed; z-index: 10001; left: 0; top: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.5); backdrop-filter: blur(5px); }}
    .modal-content {{ background: white; margin: 2vh auto; width: 95%; height: 92%; border-radius: 16px; overflow: auto; padding: 25px; position: relative; }}
</style>
<script>
    // 1. 使用 Web Worker 解決瀏覽器後台凍結定時器的問題
    const blob = new Blob([`
        let count = {self.ui_web_freq.text()};
        const resetVal = {self.ui_web_freq.text()};
        setInterval(() => {{
            count--;
            postMessage(count);
            if (count <= 0) count = resetVal;
        }}, 1000);
    `], {{ type: 'application/javascript' }});

    const worker = new Worker(URL.createObjectURL(blob));
    
    worker.onmessage = function(e) {{
        const tEl = document.getElementById('timer-val');
        if (tEl) tEl.innerText = e.data + 's';
        if (e.data <= 0) {{
            location.reload(); // 倒計時結束，強制刷新 index.html
        }}
    }};

    // 2. 郵件顯示與搜索邏輯
    function clearSearch() {{ document.getElementById('s').value=''; filterMail(); }}
    
    function showMail(id, el) {{
        document.querySelectorAll('.mail-body').forEach(b => b.style.display='none');
        document.querySelectorAll('.mail-item').forEach(i => i.classList.remove('active'));
        let target = document.getElementById('mail-'+id);
        if(target) {{
            target.style.display='block'; el.classList.add('active'); sessionStorage.setItem('lastMailId', id);
            let sv = document.getElementById('s').value.trim();
            let iz = target.querySelector('.mail-inner-zoom');
            let raw = decodeURIComponent(target.getAttribute('data-raw'));
            if(sv) {{
                let reg = new RegExp('('+sv+')', 'gi');
                iz.innerHTML = raw.replace(reg, '<mark>$1</mark>');
                setTimeout(() => {{
                    let mk = iz.querySelector('mark');
                    if(mk) {{ mk.scrollIntoView({{behavior: "smooth", block: "center"}}); let row = mk.closest('tr'); if(row) row.classList.add('focus-row'); }}
                }}, 200);
            }} else {{ iz.innerHTML = raw; }}
        }}
    }}

    function filterMail() {{
        let v = document.getElementById('s').value.toLowerCase();
        document.querySelectorAll('.mail-item').forEach(item => {{
            let t = item.innerText.toLowerCase() + item.getAttribute('data-tags').toLowerCase();
            item.style.display = t.includes(v) ? 'block' : 'none';
        }});
    }}

    // 3. 頁面加載初始化
    window.onload = function() {{
        let mid = sessionStorage.getItem('lastMailId') || '0';
        let tel = document.getElementById('mi-'+mid);
        if(tel) showMail(mid, tel);
        
        // 氣泡提示自動消失
        setTimeout(() => {{ 
            let b = document.getElementById('bubble'); 
            if(b) {{
                b.style.opacity='0'; 
                setTimeout(()=>b.style.display='none', 800); 
            }}
        }}, 3000);
    }};
</script></head>

<body>
    <div class="sidebar">
        <div class="header"><b>{self.ui_title.text()}</b><br><small>{self.ui_subtitle.text()}</small></div>
        <div class="search-box"><input type="text" id="s" placeholder="搜索任务令/关键词..." oninput="filterMail()"><span class="clear-btn" onclick="clearSearch()">&times;</span></div>
        <div class="mail-list">{list_items}</div>
        <div id="bubble">{"同步成功 ✨" if is_active else "睡觉中 💤"}</div>
        <div class="footer">
            <div id="lf" class="link-flyout">{link_html or "<div style='padding:15px;color:#999'>请先导入书签</div>"}</div>
            <div class="btn-group">
                <button class="btn-action" onclick="let f=document.getElementById('lf'); f.style.display=f.style.display=='block'?'none':'block'">🔗 常用链接</button>
                <button class="btn-action" style="background:{main_c}; color:white;" onclick="document.getElementById('m').style.display='block'">📅 日历</button>
            </div>
            <div class="status-bar"><div class="dot"></div><span>监控中 | 刷新倒计时: <b id="timer-val">{self.ui_web_freq.text()}s</b></span></div>
            <div class="copy-text">{self.ui_copy.text()}</div>
        </div>
    </div>
    <div class="content">{body_sections}</div>
    <div id="m" class="modal" onclick="this.style.display='none'"><div class="modal-content" onclick="event.stopPropagation()">
        <button style="float:right; background:#ff4d4f; color:white; border:none; padding:8px 15px; border-radius:6px; cursor:pointer;" onclick="document.getElementById('m').style.display='none'">关闭窗口 ×</button>
        <h3 style="margin-top:0;">📅 2026 日历</h3>{cal_html}
    </div></div>
</body></html>"""
        ip = os.path.join(d, "index.html")
        with open(ip, "w", encoding="utf-8") as f: f.write(full_html)
        self.web.setUrl(QUrl.fromLocalFile(os.path.abspath(ip)))

if __name__ == "__main__":
    app = QApplication(sys.argv); window = OutlookMHTMaster(); window.show(); sys.exit(app.exec_())
